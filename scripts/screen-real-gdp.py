#!/usr/bin/env python3
"""Screen archived WDI real GDP vintages; fail closed and never publish evidence.

WDI warns that NY.GDP.MKTP.KD has historically reused the same code for different
base years and that archive pages expose only current metadata. Therefore matching
code/name/current metadata is NOT sufficient to promote a cross-vintage GDP
comparison to REAL evidence.

Promotion is additionally bound to an optional reviewed release-specific methodology
attestation. That evidence must reference the exact SHA-256 fingerprints downloaded
in the same run. Missing, malformed or hash-mismatched evidence leaves the screener
fail-closed.
"""
import hashlib, io, json, math, urllib.request, zipfile
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlsplit
from openpyxl import load_workbook

RELEASES=[
 {"vintage":"2025-01-28","url":"https://databank.worldbank.org/data/download/Archive/WDI_excel_2025_01_28.zip"},
 {"vintage":"2025-07-02","url":"https://databank.worldbank.org/data/download/Archive/WDI_excel_2025_07_02.zip"},
]
CODE="NY.GDP.MKTP.KD"; CURRENT_NAME="GDP (constant 2015 US$)"; CURRENT_UNIT="constant 2015 US$"; YEAR=2023
COUNTRIES=["DEU","USA","CHN","IND","JPN","GBR","FRA","ITA","BRA","CAN","AUS","ESP","MEX","IDN","KOR"]
OUT=Path("data/screening/real-gdp-2025.json")
METHODOLOGY_ATTESTATION=Path("data/methodology/real-gdp-2025-release-evidence.json")
ARCHIVE_WARNING_URL="https://databank.worldbank.org/databases/archives"
CURRENT_METADATA_URL="https://databank.worldbank.org/metadataglossary/world-development-indicators/series/NY.GDP.MKTP.KD"

def is_worldbank_https_url(url):
    try:
        p=urlsplit(str(url))
    except Exception:
        return False
    host=(p.hostname or "").lower().rstrip(".")
    return p.scheme=="https" and bool(host) and (host=="worldbank.org" or host.endswith(".worldbank.org")) and not p.username and not p.password

def download(url):
    if not is_worldbank_https_url(url):
        raise RuntimeError(f"Refusing non-World-Bank archive URL: {url}")
    req=urllib.request.Request(url,headers={"User-Agent":"world-discovery-engine/0.5"})
    with urllib.request.urlopen(req,timeout=180) as r:
        blob=r.read()
    if not blob: raise RuntimeError(f"Empty archive: {url}")
    return blob

def extract_xlsx(blob):
    with zipfile.ZipFile(io.BytesIO(blob)) as z:
        names=[n for n in z.namelist() if n.lower().endswith('.xlsx') and not n.startswith('__MACOSX')]
        if not names: raise RuntimeError('No XLSX in archive')
        name=sorted(names,key=lambda n:(('data' not in n.lower()),-z.getinfo(n).file_size))[0]
        return name,z.read(name)

def read(blob):
    wb=load_workbook(io.BytesIO(blob),read_only=True,data_only=True)
    ws=wb['Data'] if 'Data' in wb.sheetnames else wb[wb.sheetnames[0]]
    rows=ws.iter_rows(values_only=True); header=None
    for row in rows:
        vals=[str(v).strip() if v is not None else '' for v in row]
        if all(x in vals for x in ('Country Name','Country Code','Indicator Name','Indicator Code')): header=vals; break
    if not header: raise RuntimeError('WDI header missing')
    idx={v:i for i,v in enumerate(header)}
    if str(YEAR) not in idx: raise RuntimeError(f'{YEAR} missing')
    out={}; names=set()
    for row in rows:
        if row[idx['Indicator Code']] != CODE: continue
        indicator_name=str(row[idx['Indicator Name']]).strip(); names.add(indicator_name)
        code=str(row[idx['Country Code']]).strip(); val=row[idx[str(YEAR)]]
        if code in COUNTRIES and val is not None:
            value=float(val)
            if not math.isfinite(value) or value <= 0: raise RuntimeError(f'Invalid GDP value for {code}: {value}')
            if code in out: raise RuntimeError(f'Duplicate GDP row for {code}')
            out[code]={"country":str(row[idx['Country Name']]).strip(),"value":value}
    if len(names) != 1: raise RuntimeError(f'Ambiguous indicator names in archive: {sorted(names)}')
    return out,next(iter(names))

def validate_methodology_attestation(snapshots):
    diagnostics={
      "attestationPath":str(METHODOLOGY_ATTESTATION),
      "attestationPresent":METHODOLOGY_ATTESTATION.is_file(),
      "reviewMetadataComplete":False,
      "boundToExactArchiveHashes":False,
      "authoritativeReleaseSpecificSources":False,
      "sameBaseYearUnitAndValuation":False,
    }
    if not METHODOLOGY_ATTESTATION.is_file():
        diagnostics["reason"]="No reviewed release-specific methodology attestation is present for these exact archives."
        return False,diagnostics
    try:
        att=json.loads(METHODOLOGY_ATTESTATION.read_text(encoding='utf-8'))
        if att.get('schemaVersion') != 2: raise ValueError('schemaVersion must be 2')
        if att.get('indicatorCode') != CODE: raise ValueError('indicatorCode mismatch')
        if att.get('referenceYear') != YEAR: raise ValueError('referenceYear mismatch')
        review=att.get('review')
        if not isinstance(review,dict) or review.get('status') != 'APPROVED':
            raise ValueError('review.status must be APPROVED')
        reviewer=str(review.get('reviewedBy','')).strip()
        reviewed_at=str(review.get('reviewedAtUtc','')).strip()
        if len(reviewer) < 3 or not reviewed_at.endswith('Z'):
            raise ValueError('review metadata incomplete')
        diagnostics['reviewMetadataComplete']=True
        evidence=att.get('releases')
        if not isinstance(evidence,list) or len(evidence) != len(snapshots): raise ValueError('exactly two releases required')
        by_vintage={e.get('vintage'):e for e in evidence if isinstance(e,dict)}
        if set(by_vintage) != {s['vintage'] for s in snapshots}: raise ValueError('vintage contract mismatch')
        bases=[]; units=[]; valuations=[]; hashes_ok=True; sources_ok=True
        seen_source_urls=set()
        for s in snapshots:
            e=by_vintage[s['vintage']]
            if e.get('archiveSha256') != s['archiveSha256']: hashes_ok=False
            bases.append(e.get('baseYear')); units.append(e.get('unit')); valuations.append(e.get('valuation'))
            srcs=e.get('authoritativeSources')
            if not isinstance(srcs,list) or not srcs:
                sources_ok=False
            else:
                release_has_source=False
                for src in srcs:
                    if not isinstance(src,dict) or src.get('releaseSpecific') is not True:
                        sources_ok=False; continue
                    url=str(src.get('url','')).strip()
                    title=str(src.get('title','')).strip()
                    evidence_note=str(src.get('evidence','')).strip()
                    if not is_worldbank_https_url(url):
                        sources_ok=False; continue
                    if url in (ARCHIVE_WARNING_URL,CURRENT_METADATA_URL):
                        sources_ok=False; continue
                    if url in seen_source_urls:
                        sources_ok=False; continue
                    seen_source_urls.add(url)
                    if len(title) < 5 or len(evidence_note) < 20:
                        sources_ok=False; continue
                    release_has_source=True
                if not release_has_source:
                    sources_ok=False
        diagnostics['boundToExactArchiveHashes']=hashes_ok
        diagnostics['authoritativeReleaseSpecificSources']=sources_ok
        diagnostics['sameBaseYearUnitAndValuation']=(len(set(bases))==1 and bases[0]==2015 and len(set(units))==1 and units[0]==CURRENT_UNIT and len(set(valuations))==1 and bool(valuations[0]))
        verified=all((diagnostics['reviewMetadataComplete'],diagnostics['boundToExactArchiveHashes'],diagnostics['authoritativeReleaseSpecificSources'],diagnostics['sameBaseYearUnitAndValuation']))
        diagnostics['reason']='Release-specific methodology attestation satisfies the reviewed source, exact archive-hash, base-year, unit and valuation contract.' if verified else 'Release-specific methodology attestation does not satisfy every required contract check.'
        return verified,diagnostics
    except Exception as e:
        diagnostics['reason']=f"Methodology attestation rejected: {e}"
        return False,diagnostics

def main():
    screened_at=datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace('+00:00','Z')
    snapshots=[]
    for rel in RELEASES:
        archive=download(rel['url']); member,xlsx=extract_xlsx(archive); data,name=read(xlsx)
        snapshots.append({**rel,"archiveSha256":hashlib.sha256(archive).hexdigest(),"archiveBytes":len(archive),"member":member,"indicatorNameInArchive":name,"data":data})
    present_in_both=[c for c in COUNTRIES if all(c in s['data'] for s in snapshots)]
    missing=[c for c in COUNTRIES if c not in present_in_both]
    rows=[]
    for c in present_in_both:
        a=snapshots[0]['data'][c]['value']; b=snapshots[1]['data'][c]['value']; d=b-a
        rows.append({"code":c,"country":snapshots[1]['data'][c]['country'],"first":a,"latest":b,"absoluteRevision":d,"relativeRevision":d/abs(a)})

    archive_names=[s['indicatorNameInArchive'] for s in snapshots]
    name_consistent=len(set(archive_names)) == 1
    name_matches_expected=all(n == CURRENT_NAME for n in archive_names)
    release_specific_methodology_verified,attestation=validate_methodology_attestation(snapshots)
    publishable=(not missing and name_consistent and name_matches_expected and release_specific_methodology_verified)
    result={
      "status":"SCREENING" if not publishable else "VERIFIED",
      "publishable":publishable,
      "screenedAtUtc":screened_at,
      "indicator":{"code":CODE,"currentName":CURRENT_NAME,"currentUnit":CURRENT_UNIT},
      "referenceYear":YEAR,
      "provenance":{"dataset":"World Development Indicators (WDI)","archiveWarningUrl":ARCHIVE_WARNING_URL,"currentMetadataUrl":CURRENT_METADATA_URL,"releases":[{k:s[k] for k in ('vintage','url','archiveSha256','archiveBytes','member','indicatorNameInArchive')} for s in snapshots]},
      "methodologyGate":{
        "archiveNamesConsistent":name_consistent,
        "archiveNamesMatchExpectedScreeningContract":name_matches_expected,
        "releaseSpecificBaseAndValuationVerified":release_specific_methodology_verified,
        "releaseEvidenceAttestation":attestation,
        "reason":attestation.get('reason')
      },
      "coverage":{"requested":len(COUNTRIES),"requestedCountryCodes":COUNTRIES,"rowsPresentInBothVintages":len(present_in_both),"missing":missing},
      "rows":rows,
      "promotionGate":"Do not generate public REAL GDP evidence until a reviewed release-specific methodology attestation (schema 2, APPROVED review metadata) proves identical 2015 base year, unit and valuation for both vintages, cites distinct authoritative release-specific World Bank HTTPS sources with explicit evidence notes, and binds those claims to the exact archive SHA-256 fingerprints observed by this run."
    }
    OUT.parent.mkdir(parents=True,exist_ok=True); OUT.write_text(json.dumps(result,indent=2)+"\n",encoding='utf-8')
    print(json.dumps(result,indent=2))
    if missing: raise SystemExit(f'Fail closed: missing countries {missing}')
    if not name_consistent: raise SystemExit('Fail closed: archived indicator names differ')
    if not name_matches_expected: raise SystemExit(f'Fail closed: archived indicator name differs from expected contract {CURRENT_NAME!r}')
    if not release_specific_methodology_verified: raise SystemExit('Fail closed: GDP release-specific methodology is not verified and bound to the exact archive hashes; screening output only')
if __name__=='__main__': main()
