#!/usr/bin/env python3
import csv, io, json, re, urllib.request, zipfile
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook

RELEASES=[{"vintage":"2025-01-28","url":"https://databank.worldbank.org/data/download/Archive/WDI_excel_2025_01_28.zip"},{"vintage":"2025-07-02","url":"https://databank.worldbank.org/data/download/Archive/WDI_excel_2025_07_02.zip"}]
INDICATOR_CODE="SP.POP.TOTL"; INDICATOR_NAME="Population, total"; REFERENCE_YEAR=2023; UNIT="people"
METHODOLOGY_VERSION="wdi-population-estimates-2024-note-v1"
LICENSE_URL="https://creativecommons.org/licenses/by/4.0/"
ENTITY_CODES=["DEU","USA","CHN","IND","JPN","GBR","FRA","ITA","BRA","CAN","AUS","ESP","MEX","IDN","KOR"]
ROOT=Path("site/evidence")

def download(url):
    req=urllib.request.Request(url,headers={"User-Agent":"world-discovery-engine/0.2"})
    with urllib.request.urlopen(req,timeout=180) as r:return r.read()

def extract_data_xlsx(blob):
    with zipfile.ZipFile(io.BytesIO(blob)) as z:
        names=[n for n in z.namelist() if n.lower().endswith('.xlsx') and not n.startswith('__MACOSX')]
        if not names: raise RuntimeError('No XLSX file found in WDI archive')
        preferred=sorted(names,key=lambda n:(('data' not in n.lower()),-z.getinfo(n).file_size))[0]
        return z.read(preferred),preferred

def read_indicator_rows(xlsx_bytes):
    wb=load_workbook(io.BytesIO(xlsx_bytes),read_only=True,data_only=True); ws=wb['Data'] if 'Data' in wb.sheetnames else wb[wb.sheetnames[0]]
    rows=ws.iter_rows(values_only=True); header=None
    for row in rows:
        vals=[str(v).strip() if v is not None else '' for v in row]
        if 'Country Name' in vals and 'Country Code' in vals and 'Indicator Code' in vals: header=vals; break
    if not header: raise RuntimeError('Could not locate WDI header')
    idx={name:i for i,name in enumerate(header)}; year=str(REFERENCE_YEAR); out={}; archived_names=set()
    if year not in idx: raise RuntimeError(f'Reference year {REFERENCE_YEAR} not found')
    if 'Indicator Name' not in idx: raise RuntimeError('Indicator Name column not found')
    for row in rows:
        if row[idx['Indicator Code']] != INDICATOR_CODE: continue
        code=row[idx['Country Code']]; value=row[idx[year]]
        if code in ENTITY_CODES and value is not None:
            archived_name=str(row[idx['Indicator Name']]).strip(); archived_names.add(archived_name)
            country_name=str(row[idx['Country Name']]).strip()
            if not country_name: raise RuntimeError(f'Missing archived country name for {code}')
            try: out[str(code)]={"country":country_name,"value":float(value)}
            except (TypeError,ValueError): pass
    missing=set(ENTITY_CODES)-set(out)
    if missing: raise RuntimeError(f'Missing countries: {sorted(missing)}')
    if archived_names != {INDICATOR_NAME}: raise RuntimeError(f'Unexpected archived indicator identity for {INDICATOR_CODE}: {sorted(archived_names)}')
    return out,next(iter(archived_names))

def slug(name): return re.sub(r'[^a-z0-9]+','-',name.lower()).strip('-')
def fmt(v): return f"{v:,.0f}"
def possessive(name): return f"{name}'" if name.endswith('s') else f"{name}'s"
def finding_type(delta): return 'NO_REVISION' if delta==0 else ('UPWARD_REVISION' if delta>0 else 'DOWNWARD_REVISION')

def main():
    snapshots=[]
    for rel in RELEASES:
        blob=download(rel['url']); xlsx,name=extract_data_xlsx(blob); data,archived_indicator_name=read_indicator_rows(xlsx)
        snapshots.append({**rel,"archive_file":name,"archive_sha256":__import__('hashlib').sha256(blob).hexdigest(),"archive_bytes":len(blob),"archived_indicator_name":archived_indicator_name,"data":data})
        print(f"Loaded {INDICATOR_CODE} ({archived_indicator_name}) for {len(data)} countries from {rel['vintage']}")
    for code in ENTITY_CODES:
        names={s['data'][code]['country'] for s in snapshots}
        if len(names)!=1: raise RuntimeError(f'Archived country identity changed for {code}: {sorted(names)}')
    generated=datetime.now(timezone.utc).isoformat(); manifest=[]
    for code in ENTITY_CODES:
        first=snapshots[0]['data'][code]['value']; latest=snapshots[1]['data'][code]['value']; country=snapshots[1]['data'][code]['country']
        absolute=latest-first; relative=absolute/abs(first) if first else 0.0; finding=finding_type(absolute)
        page_slug=f"{slug(country)}-population-revision-2025"; out=ROOT/page_slug; out.mkdir(parents=True,exist_ok=True)
        evidence={"schemaVersion":"1.4","status":"REAL","finding":{"type":finding,"isRevision":absolute!=0},"indicator":{"code":INDICATOR_CODE,"name":INDICATOR_NAME,"unit":UNIT,"methodologyVersion":METHODOLOGY_VERSION},"entity":{"code":code,"name":country,"entityType":"country","archivedNameVerifiedAcrossVintages":True},"referenceYear":REFERENCE_YEAR,
                  "first":{"vintage":RELEASES[0]['vintage'],"value":first,"sourceUrl":RELEASES[0]['url'],"archiveFile":snapshots[0]['archive_file'],"archiveSha256":snapshots[0]['archive_sha256'],"archiveBytes":snapshots[0]['archive_bytes'],"archivedIndicatorName":snapshots[0]['archived_indicator_name'],"archivedCountryName":snapshots[0]['data'][code]['country']},
                  "latest":{"vintage":RELEASES[1]['vintage'],"value":latest,"sourceUrl":RELEASES[1]['url'],"archiveFile":snapshots[1]['archive_file'],"archiveSha256":snapshots[1]['archive_sha256'],"archiveBytes":snapshots[1]['archive_bytes'],"archivedIndicatorName":snapshots[1]['archived_indicator_name'],"archivedCountryName":snapshots[1]['data'][code]['country']},
                  "revision":{"absolute":absolute,"relative":relative},"generatedAt":generated,"methodologyNote":"Same WDI indicator code, archived indicator name, archived country identity, unit, sovereign country and reference year across two archived 2025 releases. Regional and income aggregates are excluded.","license":"CC BY 4.0 (World Bank WDI; preserve attribution)","licenseUrl":LICENSE_URL}
        (out/'evidence.json').write_text(json.dumps(evidence,indent=2)+"\n",encoding='utf-8')
        with (out/'evidence.csv').open('w',newline='',encoding='utf-8') as f:
            w=csv.writer(f); w.writerow(['entity_code','entity','archived_country_name','indicator_code','archived_indicator_name','reference_year','vintage','value','unit','source_url','archive_file','archive_sha256','archive_bytes'])
            for key in ('first','latest'):
                x=evidence[key]; w.writerow([code,country,x['archivedCountryName'],INDICATOR_CODE,x['archivedIndicatorName'],REFERENCE_YEAR,x['vintage'],x['value'],UNIT,x['sourceUrl'],x['archiveFile'],x['archiveSha256'],x['archiveBytes']])
        unchanged=(absolute==0); direction='increased' if absolute>0 else ('decreased' if absolute<0 else 'was unchanged')
        canonical=f"https://worlddiscoverydata.com/evidence/{page_slug}/"
        description=(f"Verified World Bank WDI vintage comparison showing that the published {REFERENCE_YEAR} population estimate for {country} was unchanged between January and July 2025." if unchanged else f"Verified World Bank WDI vintage comparison showing how the published {REFERENCE_YEAR} population estimate for {country} changed between January and July 2025.")
        structured={"@context":"https://schema.org","@type":"Dataset","name":f"{country} population revision 2025","description":f"Verified comparison of the published {REFERENCE_YEAR} {country} population estimate across the 28 January 2025 and 2 July 2025 World Bank WDI archive releases.","url":canonical,"identifier":f"{INDICATOR_CODE}-{code}-{REFERENCE_YEAR}-{RELEASES[0]['vintage']}-{RELEASES[1]['vintage']}","creator":{"@type":"Organization","name":"World Discovery Engine"},"license":LICENSE_URL,"isBasedOn":[RELEASES[0]['url'],RELEASES[1]['url']],"spatialCoverage":{"@type":"Place","name":country},"temporalCoverage":f"{RELEASES[0]['vintage']}/{RELEASES[1]['vintage']}","variableMeasured":{"@type":"PropertyValue","name":INDICATOR_NAME,"propertyID":INDICATOR_CODE,"unitText":UNIT},"distribution":[{"@type":"DataDownload","encodingFormat":"application/json","contentUrl":canonical+"evidence.json"},{"@type":"DataDownload","encodingFormat":"text/csv","contentUrl":canonical+"evidence.csv"}]}
        country_possessive=possessive(country); revision_percent="0.0000%" if unchanged else f"{relative*100:+.4f}%"; revision_people="0 people" if unchanged else f"{absolute:+,.0f} people"
        what_changed=(f"The published value for {country_possessive} {REFERENCE_YEAR} population was identical in both archived releases: <strong>{fmt(first)} people</strong>. This verified null finding means no revision was observed in this vintage pair; it is not a statement about population growth between January and July 2025." if unchanged else f"The value published for {country_possessive} {REFERENCE_YEAR} population changed by <strong>{absolute:+,.0f} people</strong> ({relative*100:+.4f}%) between the two archived releases. This is a revision of a published estimate, not population growth between January and July 2025.")
        provenance=f"Indicator: {INDICATOR_NAME} ({INDICATOR_CODE}) · Archived series name verified in both releases: {INDICATOR_NAME} · Archived country identity verified in both releases: {country} ({code}) · Unit: {UNIT} · Reference year: {REFERENCE_YEAR} · Finding type: {finding} · Methodology gate: {METHODOLOGY_VERSION}. Regional aggregates are excluded. Source license: CC BY 4.0. Archive fingerprints: Jan {snapshots[0]['archive_sha256'][:12]}… · Jul {snapshots[1]['archive_sha256'][:12]}…; exact hashes, archived country/series names and archive members are included in JSON/CSV."
        html=f'''<!doctype html><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{country} population revision 2025 — World Discovery Engine</title><meta name="description" content="{description}"><link rel="canonical" href="{canonical}"><script type="application/ld+json">{json.dumps(structured,separators=(',',':'))}</script><link rel="stylesheet" href="../../styles.css"></head><body><header class="topbar"><div class="wrap"><div class="brand">World Discovery Engine</div><nav class="nav"><a href="../../index.html">Home</a><a href="../index.html">Evidence</a><a href="../../methodology/index.html">Methodology</a></nav></div></header><main><section class="evidence-head"><div class="wrap"><span class="pill">REAL WDI EVIDENCE · VERIFIED VINTAGE COMPARISON</span><h1>{country_possessive} published {REFERENCE_YEAR} population estimate {direction} between two 2025 WDI releases.</h1><p class="muted">A reproducible comparison of official archived World Development Indicators releases.</p></div></section><section class="section"><div class="wrap"><div class="facts"><div class="fact"><div class="label">28 Jan 2025</div><div class="value">{fmt(first)}</div><div class="muted">people</div></div><div class="fact"><div class="label">2 Jul 2025</div><div class="value">{fmt(latest)}</div><div class="muted">people</div></div><div class="fact"><div class="label">Revision</div><div class="value">{revision_percent}</div><div class="muted">{revision_people}</div></div></div><h2>What changed?</h2><p>{what_changed}</p><h2>Provenance</h2><div class="sourcebox">{provenance}</div><p><a href="./evidence.json">JSON evidence</a> · <a href="./evidence.csv">CSV evidence</a> · <a href="{RELEASES[0]['url']}">Jan archive</a> · <a href="{RELEASES[1]['url']}">Jul archive</a></p></div></section></main><footer class="footer"><div class="wrap">World Discovery Engine · Verified archived WDI evidence</div></footer></body></html>'''
        (out/'index.html').write_text(html,encoding='utf-8')
        manifest.append({"code":code,"country":country,"archivedCountryName":country,"archivedNameVerifiedAcrossVintages":True,"slug":page_slug,"finding":{"type":finding,"isRevision":absolute!=0},"absolute":absolute,"relative":relative,"first":{"vintage":RELEASES[0]['vintage'],"value":first},"latest":{"vintage":RELEASES[1]['vintage'],"value":latest}})
    (ROOT/'real-wdi-population-manifest.json').write_text(json.dumps({"schemaVersion":"1.4","generatedAt":generated,"indicator":{"code":INDICATOR_CODE,"name":INDICATOR_NAME,"referenceYear":REFERENCE_YEAR,"unit":UNIT},"archives":[{"vintage":s["vintage"],"url":s["url"],"archiveFile":s["archive_file"],"archiveSha256":s["archive_sha256"],"archiveBytes":s["archive_bytes"],"archivedIndicatorName":s["archived_indicator_name"]} for s in snapshots],"count":len(manifest),"items":manifest},indent=2)+"\n",encoding='utf-8')
    print(json.dumps(manifest,indent=2))
if __name__=='__main__': main()